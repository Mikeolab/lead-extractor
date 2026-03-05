"""
Export Module - Export leads to CSV and Excel formats.
"""
import csv
import io
import json
import logging
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


def leads_to_csv(leads: List[Dict[str, Any]]) -> str:
    """
    Export leads to CSV format.
    
    Returns CSV content as a string.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Business Name",
        "Emails",
        "Contact Names",
        "Phone Numbers",
        "Source URL",
        "Search Keyword",
    ])
    
    # Data
    for lead in leads:
        emails = lead.get("emails", [])
        if isinstance(emails, str):
            emails = json.loads(emails) if emails else []
        
        names = lead.get("contact_names", [])
        if isinstance(names, str):
            names = json.loads(names) if names else []
        
        phones = lead.get("phones", [])
        if isinstance(phones, str):
            phones = json.loads(phones) if phones else []
        
        writer.writerow([
            lead.get("business_name", ""),
            "; ".join(emails) if emails else "",
            "; ".join(names) if names else "",
            "; ".join(phones) if phones else "",
            lead.get("source_url", ""),
            lead.get("keyword", ""),
        ])
    
    return output.getvalue()


def leads_to_excel(leads: List[Dict[str, Any]]) -> bytes:
    """
    Export leads to Excel format with professional styling.
    
    Returns Excel file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Leads"
    
    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    alt_fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
    
    # Headers
    headers = [
        "Business Name",
        "Emails",
        "Contact Names",
        "Phone Numbers",
        "Source URL",
        "Search Keyword",
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    for row_idx, lead in enumerate(leads, 2):
        emails = lead.get("emails", [])
        if isinstance(emails, str):
            emails = json.loads(emails) if emails else []
        
        names = lead.get("contact_names", [])
        if isinstance(names, str):
            names = json.loads(names) if names else []
        
        phones = lead.get("phones", [])
        if isinstance(phones, str):
            phones = json.loads(phones) if phones else []
        
        row_data = [
            lead.get("business_name", ""),
            "\n".join(emails) if emails else "",
            "\n".join(names) if names else "",
            "\n".join(phones) if phones else "",
            lead.get("source_url", ""),
            lead.get("keyword", ""),
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = alt_fill
    
    # Column widths
    column_widths = [30, 35, 25, 20, 50, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:F{len(leads) + 1}"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

